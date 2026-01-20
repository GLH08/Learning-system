"""
Document Parser Service for importing questions from various formats
Enhanced version v2 with improved tolerance for various input formats
"""
from typing import List, Dict, Any, Optional, Tuple
import re
import json
from io import BytesIO

from docx import Document
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


class ParsedQuestion:
    """Parsed question data structure"""
    def __init__(
        self,
        index: int,
        type: str = "unknown",
        content: str = "",
        options: Optional[Dict[str, str]] = None,
        answer: Optional[str] = None,
        explanation: Optional[str] = None,
        difficulty: str = "medium",
        tags: Optional[List[str]] = None,
        source: Optional[str] = None,
        parse_status: str = "success",
        parse_message: Optional[str] = None
    ):
        self.index = index
        self.type = type
        self.content = content
        self.options = options or {}
        self.answer = answer
        self.explanation = explanation
        self.difficulty = difficulty
        self.tags = tags or []
        self.source = source
        self.parse_status = parse_status
        self.parse_message = parse_message

    def to_dict(self) -> Dict[str, Any]:
        return {
            "index": self.index,
            "type": self.type,
            "content": self.content,
            "options": self.options if self.options else None,
            "answer": self.answer,
            "explanation": self.explanation,
            "difficulty": self.difficulty,
            "tags": self.tags,
            "source": self.source,
            "hasAnswer": bool(self.answer),
            "hasExplanation": bool(self.explanation),
            "parseStatus": self.parse_status,
            "parseMessage": self.parse_message
        }


class DocumentParser:
    """Document parser for various formats with enhanced tolerance"""

    # 判断题关键词
    JUDGE_KEYWORDS = ['正确', '错误', '对', '错', 'true', 'false', '是', '否', '√', '×', 'T', 'F', 'Y', 'N']
    JUDGE_KEYWORDS_STRICT = ['正确', '错误', '对', '错', '√', '×']

    def __init__(self):
        # ============ 题目编号模式 ============
        self.question_start_pattern = re.compile(
            r'^(?:'
            r'(?:[\(（\[]?\d+[\)）\]]?[.、．:\s])'
            r'|(?:第\s*\d+\s*题[.、．:\s]*)'
            r'|(?:题目\s*\d+[.、．:\s]*)'
            r'|(?:Question\s*\d+[.、．:\s]*)'
            r')',
            re.IGNORECASE
        )

        # ============ 答案匹配模式（增强版） ============
        self.answer_patterns = [
            # 标准格式: 答案：A / 【答案】AB / (答案) C / [答案]D
            re.compile(r'[\[【\(（]\s*(?:参考|正确|标准)?\s*答案\s*[\]】\)）]\s*[:：]?\s*([A-Fa-f]+)', re.IGNORECASE),
            re.compile(r'(?:参考|正确|标准)?\s*答案\s*[:：]\s*([A-Fa-f]+)', re.IGNORECASE),
            re.compile(r'Answer\s*[:：]?\s*([A-Fa-f]+)', re.IGNORECASE),
            # 判断题答案
            re.compile(r'[\[【\(（]\s*(?:参考|正确|标准)?\s*答案\s*[\]】\)）]\s*[:：]?\s*(正确|错误|对|错|√|×|T|F|Y|N|True|False|YES|NO)', re.IGNORECASE),
            re.compile(r'(?:参考|正确|标准)?\s*答案\s*[:：]\s*(正确|错误|对|错|√|×|T|F|Y|N|True|False|YES|NO)', re.IGNORECASE),
            # 简洁格式: 答案A / 答案:A（紧凑）
            re.compile(r'^答案\s*[:：]?\s*([A-Fa-f]+)\s*$', re.IGNORECASE),
            re.compile(r'^答案\s*[:：]?\s*(正确|错误|对|错|√|×|T|F|Y|N)\s*$', re.IGNORECASE),
        ]

        # ============ 解析匹配模式 ============
        self.explanation_patterns = [
            re.compile(r'[\[【\(（]\s*(?:答案)?\s*(?:解析|分析|说明|详解|解答)\s*[\]】\)）]\s*[:：]?\s*(.+)', re.IGNORECASE | re.DOTALL),
            re.compile(r'(?:解析|分析|详解|解答)\s*[:：]\s*(.+)', re.IGNORECASE | re.DOTALL),
        ]

        # ============ 题型标签模式（用于跳过） ============
        self.type_label_pattern = re.compile(
            r'^(?:'
            r'[一二三四五六七八九十百]+\s*[、.．:：\s]\s*(?:单选|多选|判断|简答|填空|选择|问答|论述|混合)?(?:格式)?(?:测试)?(?:题)?'
            r'|[\(（]?[一二三四五六七八九十]+[\)）]?\s*$'
            r'|(?:单选|多选|判断|简答|填空|选择|问答|论述)(?:题)?\s*[:：]?\s*$'
            r'|第[一二三四五六七八九十]+(?:部分|章|节)'
            r')$',
            re.IGNORECASE
        )

        # ============ 判断题独立答案行（Y/N 也支持） ============
        self.judge_answer_line_pattern = re.compile(
            r'^[\s　]*(正确|错误|对|错|√|×|T|F|True|False|YES|NO|Y|N|是|否)[\s　]*$',
            re.IGNORECASE
        )

    def _extract_options_from_line(self, line: str) -> Dict[str, str]:
        """
        从一行中提取所有选项（支持多种格式）
        核心增强：使用更智能的切分策略
        """
        options = {}
        line = line.strip()
        if not line:
            return options

        # 策略1：检测是否是一行多选项格式（通过统计选项字母数量）
        # 匹配所有可能的选项开头位置
        option_starts = list(re.finditer(
            r'(?:^|[\s　]+)[\(（\[]?([A-Fa-f])[\]\)）]?[.、．:：\s]',
            line
        ))

        if len(option_starts) >= 2:
            # 一行有多个选项，按位置切分
            for i, match in enumerate(option_starts):
                opt_key = match.group(1).upper()
                start_pos = match.end()
                # 结束位置是下一个选项的开始，或者行尾
                if i + 1 < len(option_starts):
                    end_pos = option_starts[i + 1].start()
                else:
                    end_pos = len(line)
                opt_val = line[start_pos:end_pos].strip()
                if opt_val:
                    options[opt_key] = opt_val
            return options

        # 策略2：尝试用更宽松的多选项分隔模式
        # 支持 "A.xxx  B.xxx" 或 "A xxx  B xxx" 格式（多空格分隔）
        multi_match = re.findall(
            r'[\(（\[]?([A-Fa-f])[\]\)）]?[.、．:：]?\s*([^A-Fa-f\(（\[\s][^\s]*(?:\s+[^A-Fa-f\(（\[\s][^\s]*)*)',
            line
        )
        if len(multi_match) >= 2:
            for opt_key, opt_val in multi_match:
                opt_val = opt_val.strip()
                if opt_val:
                    options[opt_key.upper()] = opt_val
            return options

        # 策略3：单选项匹配
        # 标准格式: A. A、 A） (A) [A] 等
        match = re.match(
            r'^[\s　]*[\(（\[]?([A-Fa-f])[\]\)）]?[.、．:：\s]\s*(.+)$',
            line
        )
        if match:
            options[match.group(1).upper()] = match.group(2).strip()
            return options

        # 策略4：宽松格式 - 字母后直接跟中文内容（无分隔符）
        match = re.match(r'^[\s　]*([A-Fa-f])\s*([^\sA-Fa-f].{2,})$', line)
        if match:
            opt_key = match.group(1).upper()
            opt_val = match.group(2).strip()
            # 验证：内容应该包含中文或者是合理的选项内容
            if re.search(r'[\u4e00-\u9fff]', opt_val) or len(opt_val) >= 2:
                options[opt_key] = opt_val
                return options

        return options

    def _try_match_answer(self, line: str) -> Optional[str]:
        """尝试匹配答案"""
        line = line.strip()
        for pattern in self.answer_patterns:
            match = pattern.search(line)
            if match:
                ans = match.group(1).strip()
                # 规范化判断题答案
                ans_lower = ans.lower()
                if ans_lower in ['正确', '对', '√', 't', 'true', 'yes', 'y', '是']:
                    return '正确'
                elif ans_lower in ['错误', '错', '×', 'f', 'false', 'no', 'n', '否']:
                    return '错误'
                return ans.upper()
        return None

    def _try_match_explanation(self, line: str) -> Optional[str]:
        """尝试匹配解析"""
        line = line.strip()
        for pattern in self.explanation_patterns:
            match = pattern.search(line)
            if match:
                return match.group(1).strip()
        return None

    def _is_judge_answer_line(self, line: str) -> Optional[str]:
        """检查是否是判断题的独立答案行"""
        match = self.judge_answer_line_pattern.match(line.strip())
        if match:
            ans = match.group(1).strip()
            ans_lower = ans.lower()
            if ans_lower in ['正确', '对', '√', 't', 'true', 'yes', 'y', '是']:
                return '正确'
            elif ans_lower in ['错误', '错', '×', 'f', 'false', 'no', 'n', '否']:
                return '错误'
        return None

    def _is_question_start(self, line: str) -> bool:
        """检查是否是题目开始行"""
        return bool(self.question_start_pattern.match(line.strip()))

    def _remove_question_number(self, line: str) -> str:
        """移除题目编号，返回纯题干"""
        line = line.strip()
        result = re.sub(
            r'^(?:'
            r'[\(（\[]?\d+[\)）\]]?[.、．:\s]+'
            r'|第\s*\d+\s*题[.、．:\s]*'
            r'|题目\s*\d+[.、．:\s]*'
            r'|Question\s*\d+[.、．:\s]*'
            r')',
            '', line, flags=re.IGNORECASE
        ).strip()
        return result

    def _should_skip_line(self, line: str) -> bool:
        """检查是否应该跳过该行"""
        line = line.strip()
        if not line:
            return True
        if self.type_label_pattern.match(line):
            return True
        return False

    def _is_likely_answer_line(self, line: str) -> bool:
        """检查该行是否可能包含答案（用于提前检测）"""
        line = line.strip().lower()
        return bool(re.search(r'答案|answer', line))

    def _is_likely_explanation_line(self, line: str) -> bool:
        """检查该行是否可能包含解析"""
        line = line.strip().lower()
        return bool(re.search(r'解析|分析|详解|说明|explanation', line))

    def _parse_lines(self, lines: List[str]) -> List[ParsedQuestion]:
        """通用的行解析逻辑"""
        questions = []
        current_question = None
        question_index = 0
        pending_judge_options = []

        for line_num, line in enumerate(lines):
            line = line.strip()

            # 跳过空行和题型标签
            if self._should_skip_line(line):
                continue

            # 检查是否是新题目开头
            if self._is_question_start(line):
                # 处理上一题的判断题选项
                if current_question and pending_judge_options:
                    self._process_pending_judge_options(current_question, pending_judge_options)
                    pending_judge_options = []

                # 保存上一题
                if current_question:
                    questions.append(current_question)

                # 开始新题目
                question_index += 1
                content = self._remove_question_number(line)
                current_question = ParsedQuestion(
                    index=question_index,
                    content=content
                )
                continue

            if not current_question:
                continue

            # 优先检查答案行（因为答案行可能包含选项字母如 "答案：A"）
            if self._is_likely_answer_line(line):
                answer = self._try_match_answer(line)
                if answer:
                    current_question.answer = answer
                    continue

            # 检查解析行
            if self._is_likely_explanation_line(line):
                explanation = self._try_match_explanation(line)
                if explanation:
                    current_question.explanation = explanation
                    continue

            # 尝试提取选项
            options = self._extract_options_from_line(line)
            if options:
                current_question.options.update(options)
                continue

            # 检查是否是判断题的独立答案行
            judge_ans = self._is_judge_answer_line(line)
            if judge_ans:
                pending_judge_options.append((line_num, judge_ans, line))
                continue

            # 尝试匹配答案（非明显答案行）
            answer = self._try_match_answer(line)
            if answer:
                current_question.answer = answer
                continue

            # 尝试匹配解析（非明显解析行）
            explanation = self._try_match_explanation(line)
            if explanation:
                current_question.explanation = explanation
                continue

            # 都不匹配，追加到题目内容（仅当还没有选项时）
            if not current_question.options and not pending_judge_options:
                current_question.content += " " + line

        # 处理最后一题
        if current_question:
            if pending_judge_options:
                self._process_pending_judge_options(current_question, pending_judge_options)
            questions.append(current_question)

        # 确定题型和验证
        for q in questions:
            q.type = self._determine_type(q)
            self._validate_question(q)

        return questions

    def _process_pending_judge_options(self, question: ParsedQuestion, pending_options: List[Tuple[int, str, str]]):
        """处理暂存的判断题选项"""
        if len(pending_options) == 2:
            # 两个选项，作为判断题的 A/B 选项
            for idx, (line_num, normalized, original) in enumerate(pending_options):
                opt_key = chr(ord('A') + idx)
                question.options[opt_key] = normalized
        elif len(pending_options) == 1:
            # 只有一个，作为答案
            _, normalized, _ = pending_options[0]
            if not question.answer:
                question.answer = normalized

    def _determine_type(self, question: ParsedQuestion) -> str:
        """根据内容和选项判断题型"""
        # 有选项的情况
        if question.options:
            option_count = len(question.options)
            option_values = [str(v).lower() for v in question.options.values()]
            option_text = ' '.join(option_values)

            # 判断题特征：2个选项且包含对/错关键词
            if option_count == 2:
                if any(k in option_text for k in self.JUDGE_KEYWORDS):
                    return 'judge'

            # 根据答案判断单选/多选
            if question.answer:
                answer_clean = re.sub(r'[^A-Z]', '', question.answer.upper())
                if len(answer_clean) > 1:
                    return 'multiple'
                elif len(answer_clean) == 1:
                    return 'single'

            # 选项数量启发式
            if option_count > 4:
                return 'multiple' if option_count > 5 else 'single'

            return 'single'

        # 无选项的情况
        if question.answer:
            ans = question.answer.strip()
            if ans in self.JUDGE_KEYWORDS_STRICT or ans.upper() in ['T', 'F', 'TRUE', 'FALSE', 'Y', 'N', 'YES', 'NO', '正确', '错误']:
                return 'judge'
            if re.match(r'^[A-F]+$', ans.upper()):
                return 'single' if len(ans) == 1 else 'multiple'

        # 题干内容启发式
        content = question.content
        if re.search(r'[（\(]\s*[)）]\s*$', content):
            return 'judge'
        if re.search(r'[（\(]\s*[)）]', content):
            return 'single'

        return 'essay'

    def _validate_question(self, question: ParsedQuestion):
        """验证题目完整性"""
        warnings = []

        if not question.content or len(question.content.strip()) < 2:
            question.parse_status = 'error'
            question.parse_message = '题目内容为空或过短'
            return

        if question.type in ['single', 'multiple']:
            if not question.options:
                warnings.append('缺少选项')
            elif len(question.options) < 2:
                warnings.append('选项数量不足')

        if not question.answer:
            warnings.append('缺少答案')

        if warnings:
            question.parse_status = 'warning'
            question.parse_message = '；'.join(warnings)

    def parse_txt(self, file_content: bytes) -> List[ParsedQuestion]:
        """Parse text file (.txt)"""
        try:
            text = None
            for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-16', 'latin-1']:
                try:
                    text = file_content.decode(encoding)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue

            if text is None:
                raise Exception("无法识别文件编码")

            lines = text.split('\n')
            return self._parse_lines(lines)

        except Exception as e:
            logger.error(f"Failed to parse text file: {e}")
            raise Exception(f"文本文件解析失败: {str(e)}")

    def parse_word(self, file_content: bytes) -> List[ParsedQuestion]:
        """Parse Word document (.docx)"""
        try:
            doc = Document(BytesIO(file_content))
            lines = [para.text for para in doc.paragraphs]
            return self._parse_lines(lines)

        except Exception as e:
            logger.error(f"Failed to parse Word document: {e}")
            raise Exception(f"Word文档解析失败: {str(e)}")

    def parse_excel(self, file_content: bytes) -> List[ParsedQuestion]:
        """Parse Excel document (.xlsx)"""
        try:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active
            questions = []

            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).lower() if cell.value else "")

            col_map = self._map_excel_columns(headers)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if not any(row):
                    continue

                question = ParsedQuestion(index=row_idx)

                if col_map.get('content') is not None:
                    question.content = str(row[col_map['content']]) if row[col_map['content']] else ""

                if col_map.get('type') is not None:
                    question.type = self._normalize_type(str(row[col_map['type']])) if row[col_map['type']] else "unknown"

                if col_map.get('difficulty') is not None:
                    question.difficulty = self._normalize_difficulty(str(row[col_map['difficulty']])) if row[col_map['difficulty']] else "medium"

                if col_map.get('options') is not None:
                    options_str = str(row[col_map['options']]) if row[col_map['options']] else ""
                    question.options = self._parse_options_string(options_str)
                else:
                    for opt_key in ['A', 'B', 'C', 'D', 'E', 'F']:
                        col_key = f'option_{opt_key}'
                        if col_map.get(col_key) is not None:
                            opt_value = str(row[col_map[col_key]]) if row[col_map[col_key]] else ""
                            if opt_value and opt_value.lower() != 'none':
                                question.options[opt_key] = opt_value

                if col_map.get('answer') is not None:
                    ans = row[col_map['answer']]
                    question.answer = str(ans).strip().upper() if ans else None

                if col_map.get('explanation') is not None:
                    exp = row[col_map['explanation']]
                    question.explanation = str(exp).strip() if exp and str(exp).lower() != 'none' else None

                if col_map.get('source') is not None:
                    src = row[col_map['source']]
                    question.source = str(src).strip() if src and str(src).lower() != 'none' else None

                if question.type == "unknown":
                    question.type = self._determine_type(question)

                self._validate_question(question)
                questions.append(question)

            return questions

        except Exception as e:
            logger.error(f"Failed to parse Excel document: {e}")
            raise Exception(f"Excel文档解析失败: {str(e)}")

    def _map_excel_columns(self, headers: List[str]) -> Dict[str, int]:
        """Map Excel column headers to field names"""
        col_map = {}

        for idx, header in enumerate(headers):
            header_lower = header.lower().strip()

            if any(k in header_lower for k in ['题目', '内容', 'content', 'question', '题干', '问题']):
                col_map['content'] = idx
            elif any(k in header_lower for k in ['题型', 'type', '类型']):
                col_map['type'] = idx
            elif any(k in header_lower for k in ['难度', 'difficulty', '级别', '难易']):
                col_map['difficulty'] = idx
            elif any(k in header_lower for k in ['选项', 'options']) and header_lower not in ['选项a', '选项b', '选项c', '选项d', '选项e', '选项f']:
                col_map['options'] = idx
            elif header_lower in ['a', '选项a', 'option a', 'optiona', 'a选项']:
                col_map['option_A'] = idx
            elif header_lower in ['b', '选项b', 'option b', 'optionb', 'b选项']:
                col_map['option_B'] = idx
            elif header_lower in ['c', '选项c', 'option c', 'optionc', 'c选项']:
                col_map['option_C'] = idx
            elif header_lower in ['d', '选项d', 'option d', 'optiond', 'd选项']:
                col_map['option_D'] = idx
            elif header_lower in ['e', '选项e', 'option e', 'optione', 'e选项']:
                col_map['option_E'] = idx
            elif header_lower in ['f', '选项f', 'option f', 'optionf', 'f选项']:
                col_map['option_F'] = idx
            elif any(k in header_lower for k in ['答案', 'answer', '正确答案', '正确选项']):
                col_map['answer'] = idx
            elif any(k in header_lower for k in ['解析', 'explanation', '说明', '详解', '分析', '解答']):
                col_map['explanation'] = idx
            elif any(k in header_lower for k in ['来源', 'source', '出处', '出自']):
                col_map['source'] = idx

        return col_map

    def _parse_options_string(self, options_str: str) -> Dict[str, str]:
        """Parse options from a single string"""
        options = {}
        if not options_str:
            return options

        parts = re.split(r'[;；\n\r]', options_str)

        for part in parts:
            part = part.strip()
            if not part:
                continue
            extracted = self._extract_options_from_line(part)
            options.update(extracted)

        if not options:
            options = self._extract_options_from_line(options_str)

        return options

    def _normalize_type(self, type_str: str) -> str:
        """Normalize question type string"""
        if not type_str:
            return 'unknown'
        type_lower = type_str.lower()

        if any(k in type_lower for k in ['单选', 'single', '单项', '选择题']):
            return 'single'
        elif any(k in type_lower for k in ['多选', 'multiple', '多项', '不定项']):
            return 'multiple'
        elif any(k in type_lower for k in ['判断', 'judge', 'true/false', 'tf', '对错', '是非']):
            return 'judge'
        elif any(k in type_lower for k in ['简答', 'essay', '问答', '简述', '论述', '解答']):
            return 'essay'
        elif any(k in type_lower for k in ['填空', 'blank', 'fill']):
            return 'essay'

        return 'unknown'

    def _normalize_difficulty(self, diff_str: str) -> str:
        """Normalize difficulty string"""
        if not diff_str:
            return 'medium'
        diff_lower = diff_str.lower()

        if any(k in diff_lower for k in ['简单', 'easy', '容易', '低', '初级', '基础']):
            return 'easy'
        elif any(k in diff_lower for k in ['困难', 'hard', '难', '高', '高级', '较难']):
            return 'hard'
        else:
            return 'medium'


def get_document_parser() -> DocumentParser:
    """Get document parser instance"""
    return DocumentParser()
